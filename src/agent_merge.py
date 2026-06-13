import json
import os
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.config import DEFAULT_API_URL, DEFAULT_OCR_MODEL

def merge_files_to_excel(task_description, api_key, api_url=None, model_name=None,
                         file_path_1="", file_path_2="", file_path_3="",
                         output_path="consolidado.xlsx", chat_callback=None, status_callback=None):
    """
    Consolida dados de até 3 arquivos com base nas instruções e cria uma planilha Excel estilizada profissionalmente.
    """
    url = api_url or DEFAULT_API_URL
    model = model_name or DEFAULT_OCR_MODEL
    
    def log_chat(sender, text):
        if chat_callback:
            chat_callback(sender, text)
            
    def log_status(text):
        if status_callback:
            status_callback(text)

    log_status("Status: Lendo arquivos de entrada...")
    
    # Carrega arquivos
    content1 = ""
    if file_path_1 and os.path.exists(file_path_1):
        with open(file_path_1, 'r', encoding='utf-8', errors='ignore') as f:
            content1 = f.read()

    content2 = ""
    if file_path_2 and os.path.exists(file_path_2):
        with open(file_path_2, 'r', encoding='utf-8', errors='ignore') as f:
            content2 = f.read()

    content3 = ""
    if file_path_3 and os.path.exists(file_path_3):
        with open(file_path_3, 'r', encoding='utf-8', errors='ignore') as f:
            content3 = f.read()

    if not content1 and not content2 and not content3:
        log_chat("Sistema", "Erro: Nenhum arquivo de entrada válido foi selecionado ou os arquivos estão vazios.")
        return False

    log_status("Status: Consolidadando dados com IA...")
    
    prompt = (
        f"Você é um engenheiro de dados especialista. O usuário deseja realizar a seguinte consolidação de arquivos:\n"
        f"'{task_description}'.\n\n"
        f"Aqui estão as informações dos arquivos de entrada:\n"
        f"=== CONTEÚDO ARQUIVO 1 ===\n{content1}\n==========================\n\n"
        f"=== CONTEÚDO ARQUIVO 2 ===\n{content2}\n==========================\n\n"
        f"=== CONTEÚDO ARQUIVO 3 ===\n{content3}\n==========================\n\n"
        f"Por favor, mescle e consolide os dados desses arquivos em uma única tabela estruturada organizada de forma consistente.\n"
        f"Retorne obrigatoriamente um objeto JSON puro (sem markdown) contendo a lista de linhas sob a chave 'rows', "
        f"onde cada linha é um dicionário chave-valor (cabeçalho da coluna -> valor).\n"
        f"Exemplo de resposta:\n"
        f"{{\n"
        f"  \"rows\": [\n"
        f"    {{\"Nome\": \"Willian\", \"Email\": \"willian@teste.com\", \"Origem\": \"Arquivo 1\"}},\n"
        f"    {{\"Nome\": \"Danilo\", \"Email\": \"danilo@teste.com\", \"Origem\": \"Arquivo 2\"}}\n"
        f"  ]\n"
        f"}}"
    )

    try:
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        payload = {
            "model": model,
            "response_format": {"type": "json_object"},
            "messages": [
                {"role": "user", "content": prompt}
            ],
            "max_tokens": 3000,
            "temperature": 0.2
        }
        
        res = requests.post(f"{url}/chat/completions", headers=headers, json=payload, timeout=75)
        res.raise_for_status()
        res_data = res.json()
        ai_message = res_data['choices'][0]['message']['content'].strip()

        # Sanitiza markdown
        if ai_message.startswith("```json"):
            ai_message = ai_message[7:]
        if ai_message.endswith("```"):
            ai_message = ai_message[:-3]
        ai_message = ai_message.strip()

        data_json = json.loads(ai_message)
        rows = data_json.get("rows", [])

        if not rows:
            raise ValueError("O agente de IA retornou uma lista vazia de linhas.")

        log_status("Status: Gerando planilha Excel estilizada...")
        
        # Cria a planilha do openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consolidado RPA"
        
        # Habilita grid lines explicitamente
        ws.views.sheetView[0].showGridLines = True

        # Cabeçalhos
        headers_list = list(rows[0].keys())
        for col_idx, h_text in enumerate(headers_list, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h_text)
            cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            # Tema azul escuro corporativo
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        # Linhas de Dados e Bordas
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        for row_idx, r_data in enumerate(rows, start=2):
            for col_idx, h_text in enumerate(headers_list, start=1):
                val = r_data.get(h_text, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="center")
                cell.border = thin_border
                
                # Zebrado (linhas pares com cinza claro)
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Auto-ajuste de largura das colunas
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Salva arquivo
        out_p = output_path or "consolidado.xlsx"
        wb.save(out_p)
        log_chat("Agente", f"Consolidação Excel gerada com sucesso em: '{out_p}'")
        return True

    except Exception as e:
        log_chat("Sistema", f"Erro crítico na mesclagem Excel: {str(e)}")
        print(f"[MERGE ERROR] Falha ao mesclar dados Excel: {e}")
        return False
