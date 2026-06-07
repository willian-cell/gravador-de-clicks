import threading
import time
import ctypes
try:
    # Habilita DPI awareness para Windows Vista / 7 / 8 / 10 / 11
    ctypes.windll.shcore.SetProcessDpiAwareness(2) # PROCESS_PER_MONITOR_DPI_AWARE
except Exception:
    try:
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass
import tkinter as tk
from tkinter import messagebox, simpledialog
import sqlite3
import json
import os
from pathlib import Path

try:
    from pynput import mouse
    from pynput.mouse import Controller, Button
    from pynput import keyboard
except ImportError:
    raise ImportError(
        "O pacote 'pynput' é necessário para capturar cliques do mouse fora do Tkinter. "
        "Instale com: python -m pip install pynput"
    )

try:
    from PIL import Image, ImageTk
except ImportError:
    raise ImportError(
        "O pacote 'pillow' é necessário para carregar imagens. "
        "Instale com: python -m pip install pillow"
    )

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False


class Action:
    """Representa uma ação gravada (clique, digitação, espera, etc)."""
    
    TYPE_CLICK = "CLICK"
    TYPE_TYPE = "TYPE"
    TYPE_WAIT = "WAIT"
    TYPE_READ_FILE = "READ_FILE"
    TYPE_SCROLL = "SCROLL"
    TYPE_AI = "IA"
    
    def __init__(self, action_type, **kwargs):
        self.type = action_type
        self.timestamp = kwargs.get('timestamp', time.time())
        self.x = kwargs.get('x', 0)
        self.y = kwargs.get('y', 0)
        self.button = kwargs.get('button', 'Button.left')
        self.text = kwargs.get('text', '')
        self.duration = kwargs.get('duration', 0)
        self.file_path = kwargs.get('file_path', '')
        self.read_mode = kwargs.get('read_mode', 'complete')
        self.file_selector = kwargs.get('file_selector', '')
        self.dx = kwargs.get('dx', 0)
        self.dy = kwargs.get('dy', 0)
        self.ai_task = kwargs.get('ai_task', '')
        self.ai_mode = kwargs.get('ai_mode', 'ocr')
        self.file_path_2 = kwargs.get('file_path_2', '')
        self.file_path_3 = kwargs.get('file_path_3', '')
        self.output_path = kwargs.get('output_path', '')
    
    def to_dict(self):
        """Converte ação para dicionário."""
        return {
            'type': self.type,
            'timestamp': self.timestamp,
            'x': self.x,
            'y': self.y,
            'button': self.button,
            'text': self.text,
            'duration': self.duration,
            'file_path': self.file_path,
            'read_mode': self.read_mode,
            'file_selector': self.file_selector,
            'dx': self.dx,
            'dy': self.dy,
            'ai_task': self.ai_task,
            'ai_mode': self.ai_mode,
            'file_path_2': self.file_path_2,
            'file_path_3': self.file_path_3,
            'output_path': self.output_path
        }
    
    @staticmethod
    def from_dict(data):
        """Cria ação a partir de dicionário."""
        action_type = data.get('type', Action.TYPE_CLICK)
        action = Action(action_type)
        action.timestamp = data.get('timestamp', time.time())
        action.x = data.get('x', 0)
        action.y = data.get('y', 0)
        action.button = data.get('button', 'Button.left')
        action.text = data.get('text', '')
        action.duration = data.get('duration', 0)
        action.file_path = data.get('file_path', '')
        action.read_mode = data.get('read_mode', 'complete')
        action.file_selector = data.get('file_selector', '')
        action.dx = data.get('dx', 0)
        action.dy = data.get('dy', 0)
        action.ai_task = data.get('ai_task', '')
        action.ai_mode = data.get('ai_mode', 'ocr')
        action.file_path_2 = data.get('file_path_2', '')
        action.file_path_3 = data.get('file_path_3', '')
        action.output_path = data.get('output_path', '')
        return action
    
    def __str__(self):
        """Representação legível da ação."""
        if self.type == Action.TYPE_CLICK:
            return f"Clique em ({self.x}, {self.y}) - {self.button}"
        elif self.type == Action.TYPE_TYPE:
            text_preview = self.text[:30] + "..." if len(self.text) > 30 else self.text
            return f"Digitar: {text_preview}"
        elif self.type == Action.TYPE_WAIT:
            return f"Aguardar {self.duration}s"
        elif self.type == Action.TYPE_READ_FILE:
            mode_desc = {
                'complete': 'Conteúdo Completo',
                'line': f'Linha {self.file_selector}',
                'json': f'Chave JSON: {self.file_selector}',
                'csv': f'CSV: {self.file_selector}'
            }.get(self.read_mode, self.read_mode)
            filename = os.path.basename(self.file_path) if self.file_path else "Não selecionado"
            return f"Ler arquivo ({mode_desc}): {filename}"
        elif self.type == Action.TYPE_SCROLL:
            direction = "baixo" if self.dy < 0 else "cima"
            return f"Rolar mouse: {direction} (dy: {self.dy}, dx: {self.dx})"
        elif self.type == Action.TYPE_AI:
            mode_desc = {
                'ocr': 'Visão/OCR (Desktop)',
                'selenium': 'Selenium Web',
                'excel_merge': 'Mesclar Arquivos'
            }.get(self.ai_mode, self.ai_mode)
            task_desc = self.ai_task[:25] + "..." if len(self.ai_task) > 25 else self.ai_task
            return f"IA ({mode_desc}): {task_desc or 'Sem instrução'}"
        return f"Ação desconhecida"


def get_file_content_to_type(file_path, read_mode, selector):
    """
    Lê um arquivo do disco e extrai o conteúdo a ser digitado
    com base no modo de leitura (read_mode) e no seletor (selector).
    """
    if not file_path:
        raise ValueError("Caminho do arquivo não especificado.")
        
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
        
    # Lê todo o conteúdo como string
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()
        
    if read_mode == 'complete':
        return content
        
    elif read_mode == 'line':
        lines = content.splitlines()
        try:
            line_idx = int(selector) - 1 # 1-indexado
            if 0 <= line_idx < len(lines):
                return lines[line_idx]
            else:
                raise IndexError(f"Linha {selector} fora do limite (total de linhas: {len(lines)})")
        except ValueError:
            raise ValueError("O seletor para modo 'Linha' deve ser um número inteiro.")
            
    elif read_mode == 'json':
        try:
            data = json.loads(content)
            # Suporta chaves aninhadas separadas por ponto, ex: usuario.nome
            keys = selector.split('.')
            val = data
            for k in keys:
                if isinstance(val, dict):
                    val = val[k]
                elif isinstance(val, list):
                    val = val[int(k)]
                else:
                    raise KeyError(f"Chave '{k}' não encontrada")
            return str(val)
        except Exception as e:
            raise ValueError(f"Erro ao obter chave JSON '{selector}': {e}")
            
    elif read_mode == 'csv':
        import csv
        try:
            lines = content.splitlines()
            reader = csv.reader(lines)
            rows = list(reader)
            if not rows:
                return ""
            
            # Formato de seletor 1: "linha,coluna" (ex: "1,2")
            parts = selector.split(',')
            if len(parts) == 2:
                try:
                    row_idx = int(parts[0].strip()) - 1
                    col_idx = int(parts[1].strip()) - 1
                    if 0 <= row_idx < len(rows) and 0 <= col_idx < len(rows[row_idx]):
                        return rows[row_idx][col_idx]
                    else:
                        raise IndexError(f"Coordenadas CSV {selector} fora dos limites.")
                except ValueError:
                    pass # Tenta o próximo padrão se não for numérico
            
            # Formato de seletor 2: Nome da coluna (busca a coluna na linha 2)
            col_selector = selector.strip()
            # Tenta ver se é um índice numérico simples de coluna
            if col_selector.isdigit():
                col_idx = int(col_selector) - 1
                if len(rows) > 0 and 0 <= col_idx < len(rows[0]):
                    return rows[0][col_idx]
            else:
                # Trata a primeira linha como cabeçalho
                header = rows[0]
                if col_selector in header:
                    col_idx = header.index(col_selector)
                    if len(rows) > 1:
                        return rows[1][col_idx] # primeira linha de dados
                    
            raise ValueError(f"Formato do seletor CSV inválido ou não encontrado: '{selector}'. Use 'linha,coluna' (ex: 1,2) ou nome da coluna.")
        except Exception as e:
            raise ValueError(f"Erro ao parsear CSV: {e}")
            
    return content


class DatabaseManager:
    """Gerencia persistência de gravações em banco de dados SQLite."""
    
    def __init__(self, db_path="recordings.db"):
        self.db_path = db_path
        self.init_database()
    
    def init_database(self):
        """Inicializa o banco de dados com schema."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS recordings (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL UNIQUE,
                        clicks_data TEXT NOT NULL,
                        created_at TIMESTAMP,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS settings (
                        key TEXT PRIMARY KEY,
                        value TEXT NOT NULL
                    )
                """)
                
                # Migração: Adiciona novas colunas se não existirem
                cursor.execute("PRAGMA table_info(recordings)")
                columns = [col[1] for col in cursor.fetchall()]
                if "description" not in columns:
                    cursor.execute("ALTER TABLE recordings ADD COLUMN description TEXT")
                if "last_run_at" not in columns:
                    cursor.execute("ALTER TABLE recordings ADD COLUMN last_run_at TIMESTAMP")
                if "run_count" not in columns:
                    cursor.execute("ALTER TABLE recordings ADD COLUMN run_count INTEGER DEFAULT 0")
                
                # Cria a tabela de skills
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS skills (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL UNIQUE,
                        prompt TEXT NOT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                # Semeia skills padrões caso esteja vazia
                cursor.execute("SELECT COUNT(*) FROM skills")
                if cursor.fetchone()[0] == 0:
                    default_skills = [
                        ("Disparador WhatsApp", "Você é um especialista em disparar mensagens do WhatsApp. Encontre o campo de busca de contatos, procure pelo contato correto, clique na conversa, escreva a mensagem e envie."),
                        ("Mesclador de Notas Fiscais", "Você é um especialista em consolidar dados de notas fiscais. Extraia informações relevantes como data, valor, CNPJ, cliente e crie uma planilha formatada."),
                        ("Preenchedor de Cadastro", "Você é um especialista em preenchimento de cadastros. Encontre os campos de formulário como Nome, CPF, Endereço, insira os dados correspondentes da fonte e clique em salvar.")
                    ]
                    cursor.executemany("INSERT INTO skills (name, prompt) VALUES (?, ?)", default_skills)
                    
                conn.commit()
        except Exception as e:
            print(f"Erro ao inicializar banco de dados: {e}")

    def get_setting(self, key, default=""):
        """Recupera uma configuração global do banco."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT value FROM settings WHERE key = ?", (key,))
                row = cursor.fetchone()
                return row[0] if row else default
        except Exception:
            return default

    def set_setting(self, key, value):
        """Salva ou atualiza uma configuração global no banco."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, str(value)))
                conn.commit()
            return True
        except Exception as e:
            print(f"Erro ao salvar configuração '{key}': {e}")
            return False
    
    def add_recording(self, name, clicks):
        """Adiciona nova gravação ao banco de dados."""
        try:
            # Converte Action objects para dicts
            if clicks and isinstance(clicks[0], Action):
                clicks_json = json.dumps([action.to_dict() for action in clicks])
            else:
                clicks_json = json.dumps(clicks)
            
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO recordings (name, clicks_data, created_at, run_count)
                    VALUES (?, ?, CURRENT_TIMESTAMP, 0)
                """, (name, clicks_json))
                conn.commit()
            return True
        except sqlite3.IntegrityError:
            raise ValueError(f"Gravação '{name}' já existe.")
        except Exception as e:
            print(f"Erro ao salvar gravação: {e}")
            return False
    
    def get_all_recordings(self):
        """Retorna todas as gravações do banco de dados."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT id, name, clicks_data, created_at, description, last_run_at, run_count FROM recordings ORDER BY created_at DESC")
                rows = cursor.fetchall()
                return [dict(row) for row in rows]
        except Exception as e:
            print(f"Erro ao recuperar gravações: {e}")
            return []
    
    def get_recording(self, name):
        """Retorna uma gravação específica pelo nome."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT id, name, clicks_data, created_at, description, last_run_at, run_count FROM recordings WHERE name = ?", (name,))
                row = cursor.fetchone()
                if row:
                    record = dict(row)
                    record['clicks'] = json.loads(record['clicks_data'])
                    return record
                return None
        except Exception as e:
            print(f"Erro ao recuperar gravação: {e}")
            return None
    
    def update_recording(self, name, new_name=None, clicks=None):
        """Atualiza nome ou cliques de uma gravação."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                if new_name and clicks:
                    clicks_json = json.dumps(clicks)
                    cursor.execute("""
                        UPDATE recordings 
                        SET name = ?, clicks_data = ?, updated_at = CURRENT_TIMESTAMP
                        WHERE name = ?
                    """, (new_name, clicks_json, name))
                elif new_name:
                    cursor.execute("""
                        UPDATE recordings 
                        SET name = ?, updated_at = CURRENT_TIMESTAMP
                        WHERE name = ?
                    """, (new_name, name))
                elif clicks:
                    clicks_json = json.dumps(clicks)
                    cursor.execute("""
                        UPDATE recordings 
                        SET clicks_data = ?, updated_at = CURRENT_TIMESTAMP
                        WHERE name = ?
                    """, (clicks_json, name))
                conn.commit()
            return True
        except Exception as e:
            print(f"Erro ao atualizar gravação: {e}")
            return False
    
    def delete_recording(self, name):
        """Deleta uma gravação do banco de dados."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM recordings WHERE name = ?", (name,))
                conn.commit()
            return True
        except Exception as e:
            print(f"Erro ao deletar gravação: {e}")
            return False

    def get_all_skills(self):
        """Retorna todas as skills (perfis de IA) cadastrados."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT id, name, prompt FROM skills ORDER BY name ASC")
                rows = cursor.fetchall()
                return [dict(row) for row in rows]
        except Exception as e:
            print(f"Erro ao recuperar skills: {e}")
            return []

    def add_skill(self, name, prompt):
        """Adiciona ou atualiza uma skill/perfil de IA no banco."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT OR REPLACE INTO skills (name, prompt)
                    VALUES (?, ?)
                """, (name, prompt))
                conn.commit()
            return True
        except Exception as e:
            print(f"Erro ao adicionar/atualizar skill: {e}")
            return False

    def delete_skill(self, name):
        """Deleta uma skill/perfil de IA pelo nome."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM skills WHERE name = ?", (name,))
                conn.commit()
            return True
        except Exception as e:
            print(f"Erro ao deletar skill: {e}")
            return False

    def increment_run_count(self, name):
        """Incrementa o contador de execuções de uma gravação e atualiza a data da última execução."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE recordings 
                    SET run_count = COALESCE(run_count, 0) + 1, last_run_at = CURRENT_TIMESTAMP
                    WHERE name = ?
                """, (name,))
                conn.commit()
            return True
        except Exception as e:
            print(f"Erro ao incrementar run_count para '{name}': {e}")
            return False


class ClickRecorderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gravador de Cliques com Agente IA")
        self.root.geometry("1100x660")
        self.root.resizable(True, True)

        self.db = DatabaseManager()
        self.recordings = []
        self.recording = False
        self.listener = None
        self.keyboard_listener = None
        self.current_clicks = []
        self.playing = False
        self.last_key_time = 0

        # Histórico de mensagens do chat e estados do agente
        self.chat_history_list = []
        self.user_message_event = threading.Event()
        self.last_user_message = ""
        self.agent_active = False

        self.build_ui()
        self.load_recordings()

    def build_ui(self):
        # Logo no topo
        logo_frame = tk.Frame(self.root, bg="white", height=60)
        logo_frame.pack(fill=tk.X, side=tk.TOP)
        logo_frame.pack_propagate(False)

        self.create_logo(logo_frame)

        # Frame principal
        main_frame = tk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True)

        left_frame = tk.Frame(main_frame, padx=10, pady=10)
        left_frame.pack(side=tk.LEFT, fill=tk.Y)

        tk.Label(left_frame, text="Automações Salvas", font=("Arial", 12, "bold")).pack(anchor=tk.W)
        self.recordings_listbox = tk.Listbox(left_frame, width=30, height=12)
        self.recordings_listbox.pack(pady=(5, 10))

        button_frame = tk.Frame(left_frame)
        button_frame.pack(fill=tk.X)

        self.execute_button = tk.Button(button_frame, text="Executar automação", command=self.execute_recording, width=20, bg="#ff9800", fg="white", font=("Arial", 10, "bold"))
        self.execute_button.pack(pady=5)

        self.stop_execution_button = tk.Button(button_frame, text="Parar execução", command=self.stop_execution, width=20, bg="#f44336", fg="white", font=("Arial", 10, "bold"))
        self.stop_execution_button.pack(pady=5)
        self.stop_execution_button.config(state=tk.DISABLED)

        self.new_automation_button = tk.Button(button_frame, text="Nova automação", command=self.new_automation, width=20, bg="#4caf50", fg="white", font=("Arial", 10, "bold"))
        self.new_automation_button.pack(pady=5)

        self.edit_button = tk.Button(button_frame, text="Editar automação", command=self.edit_automation, width=20, bg="#2196f3", fg="white", font=("Arial", 10, "bold"))
        self.edit_button.pack(pady=5)

        self.delete_button = tk.Button(button_frame, text="Excluir automação", command=self.delete_recording, width=20)
        self.delete_button.pack(pady=5)

        middle_frame = tk.Frame(main_frame, padx=10, pady=10)
        middle_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        tk.Label(middle_frame, text="Controle de Gravação", font=("Arial", 12, "bold")).pack(anchor=tk.W)

        self.status_var = tk.StringVar(value="Pronto para iniciar uma nova gravação.")
        self.status_label = tk.Label(middle_frame, textvariable=self.status_var, font=("Arial", 11), wraplength=380, justify=tk.LEFT)
        self.status_label.pack(pady=(5, 15), anchor=tk.W)

        self.start_button = tk.Button(middle_frame, text="Iniciar gravação", command=self.start_recording, width=22, bg="#4caf50", fg="white", font=("Arial", 11, "bold"))
        self.start_button.pack(pady=(0, 10))

        self.stop_button = tk.Button(middle_frame, text="■ Parar gravação", command=self.stop_recording, width=22, bg="#f44336", fg="white", font=("Arial", 11, "bold"))
        self.stop_button.pack(pady=(0, 20))
        self.stop_button.config(state=tk.DISABLED)

        self.current_count_var = tk.StringVar(value="Contador: 0 cliques registrados")
        tk.Label(middle_frame, textvariable=self.current_count_var, font=("Arial", 11)).pack(anchor=tk.W)

        tk.Label(middle_frame, text="Dicas:", font=("Arial", 11, "bold")).pack(anchor=tk.W, pady=(20, 5))
        tips = (
            "1. Clique em 'Iniciar gravação' para começar.\n"
            "2. Aguarde 3, 2, 1 e depois use o mouse livremente.\n"
            "3. Clique em 'Parar gravação' quando quiser finalizar.\n"
            "4. Você pode editar ou excluir registros depois."
        )
        tk.Label(middle_frame, text=tips, justify=tk.LEFT, wraplength=380, font=("Arial", 10)).pack(anchor=tk.W)

        # Seção Selenium
        if SELENIUM_AVAILABLE:
            tk.Label(middle_frame, text="Automação Web (Selenium)", font=("Arial", 11, "bold")).pack(anchor=tk.W, pady=(20, 5))
            self.selenium_button = tk.Button(middle_frame, text="Abrir Navegador Automatizado", command=self.open_selenium, width=22, bg="#9c27b0", fg="white", font=("Arial", 10, "bold"))
            self.selenium_button.pack(pady=(0, 10))

        # Configurações Globais
        config_frame = tk.LabelFrame(middle_frame, text="Configurações Globais", font=("Arial", 11, "bold"), padx=10, pady=5)
        config_frame.pack(fill=tk.X, pady=(15, 0), side=tk.BOTTOM)

        tk.Label(config_frame, text="Espera Máxima (s):", font=("Arial", 10)).grid(row=0, column=0, sticky=tk.W, pady=5)
        self.max_wait_var = tk.StringVar(value=self.db.get_setting("max_wait_delay", "10.0"))
        max_wait_entry = tk.Entry(config_frame, textvariable=self.max_wait_var, width=8)
        max_wait_entry.grid(row=0, column=1, sticky=tk.W, pady=5, padx=5)

        tk.Label(config_frame, text="OpenAI API Key:", font=("Arial", 10)).grid(row=1, column=0, sticky=tk.W, pady=5)
        self.api_key_var = tk.StringVar(value=self.db.get_setting("openai_api_key", ""))
        self.api_key_entry = tk.Entry(config_frame, textvariable=self.api_key_var, width=22, show="*")
        self.api_key_entry.grid(row=1, column=1, sticky=tk.W, pady=5, padx=5)

        def save_global_configs():
            self.db.set_setting("max_wait_delay", self.max_wait_var.get())
            self.db.set_setting("openai_api_key", self.api_key_var.get())
            messagebox.showinfo("Configurações", "Configurações globais salvas com sucesso!")

        tk.Button(config_frame, text="Salvar", command=save_global_configs, bg="#4caf50", fg="white", font=("Arial", 9, "bold"), width=8).grid(row=1, column=2, padx=10, pady=5)

        # Painel do Agente de IA (Chat) à direita
        chat_frame = tk.Frame(main_frame, padx=10, pady=10, bg="#f5f5f5", width=340)
        chat_frame.pack(side=tk.RIGHT, fill=tk.BOTH)
        chat_frame.pack_propagate(False)

        tk.Label(chat_frame, text="Painel do Agente IA", font=("Arial", 12, "bold"), bg="#f5f5f5", fg="#0097A7").pack(anchor=tk.W, pady=(0, 2))

        self.agent_status_var = tk.StringVar(value="Status: Inativo")
        self.agent_status_lbl = tk.Label(chat_frame, textvariable=self.agent_status_var, font=("Arial", 9, "italic"), bg="#f5f5f5", fg="gray")
        self.agent_status_lbl.pack(anchor=tk.W, pady=(0, 5))

        chat_history_frame = tk.Frame(chat_frame)
        chat_history_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        chat_scroll = tk.Scrollbar(chat_history_frame)
        chat_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.chat_history = tk.Text(chat_history_frame, wrap=tk.WORD, yscrollcommand=chat_scroll.set, font=("Arial", 9), state=tk.DISABLED)
        self.chat_history.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        chat_scroll.config(command=self.chat_history.yview)

        chat_input_frame = tk.Frame(chat_frame, bg="#f5f5f5")
        chat_input_frame.pack(fill=tk.X, pady=(5, 0))

        self.chat_input = tk.Entry(chat_input_frame, font=("Arial", 10))
        self.chat_input.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=3)
        self.chat_input.bind("<Return>", lambda e: self.send_user_message())

        chat_send_btn = tk.Button(chat_input_frame, text="Enviar", command=self.send_user_message, bg="#0097A7", fg="white", font=("Arial", 9, "bold"), width=8)
        chat_send_btn.pack(side=tk.RIGHT, padx=(5, 0))

    def start_recording(self):
        if self.recording:
            return

        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.DISABLED)
        self.status_var.set("Contagem regressiva: 3...")
        self.current_count_var.set("Contador: 0 cliques registrados")
        threading.Thread(target=self.countdown_and_start, daemon=True).start()

    def countdown_and_start(self):
        for count in [3, 2, 1]:
            self.status_var.set(f"Iniciando em {count}...")
            time.sleep(1)

        self.current_clicks = []
        self.recording = True
        self.last_key_time = time.time()
        self.status_var.set("Gravando cliques e digitação... pressione o botão PARA para finalizar.")
        self.stop_button.config(state=tk.NORMAL)

        # Inicia listener de mouse
        self.listener = mouse.Listener(on_click=self.on_click, on_scroll=self.on_scroll)
        self.listener.start()
        
        # Inicia listener de teclado
        self.keyboard_listener = keyboard.Listener(on_press=self.on_key_press, on_release=self.on_key_release)
        self.keyboard_listener.start()

    def on_click(self, x, y, button, pressed):
        if pressed and self.recording:
            # Filtra cliques que ocorrem dentro da própria janela do aplicativo
            try:
                win_x = self.root.winfo_rootx()
                win_y = self.root.winfo_rooty()
                win_w = self.root.winfo_width()
                win_h = self.root.winfo_height()
                if win_x <= x <= win_x + win_w and win_y <= y <= win_y + win_h:
                    # Clique dentro da janela do app (ex: no botão Parar). Ignorar.
                    return
            except Exception:
                pass

            action = Action(
                Action.TYPE_CLICK,
                timestamp=time.time(),
                x=x,
                y=y,
                button=str(button)
            )
            self.current_clicks.append(action)
            self.current_count_var.set(f"Contador: {len(self.current_clicks)} ações registradas")

    def on_scroll(self, x, y, dx, dy):
        if not self.recording:
            return
            
        # Filtra rolagens que ocorrem dentro da própria janela do aplicativo
        try:
            win_x = self.root.winfo_rootx()
            win_y = self.root.winfo_rooty()
            win_w = self.root.winfo_width()
            win_h = self.root.winfo_height()
            if win_x <= x <= win_x + win_w and win_y <= y <= win_y + win_h:
                # Rolagem dentro da janela do app. Ignorar.
                return
        except Exception:
            pass

        now = time.time()
        can_merge = False
        
        # Agrupa se a última ação gravada for rolagem e recente (menos de 0.5 segundos)
        if self.current_clicks:
            last_action = self.current_clicks[-1]
            if last_action.type == Action.TYPE_SCROLL:
                time_diff = now - last_action.timestamp
                if time_diff < 0.5:
                    can_merge = True
                    
        if can_merge:
            last_action = self.current_clicks[-1]
            last_action.dx += dx
            last_action.dy += dy
            last_action.timestamp = now
        else:
            action = Action(
                Action.TYPE_SCROLL,
                timestamp=now,
                dx=dx,
                dy=dy
            )
            self.current_clicks.append(action)
            
        self.current_count_var.set(f"Contador: {len(self.current_clicks)} ações registradas")

    def on_key_press(self, key):
        """Captura digitação de teclas e as agrupa de forma inteligente em uma única ação."""
        if not self.recording:
            return
        
        try:
            char = None
            is_backspace = False
            
            # Identifica a tecla digitada
            if hasattr(key, 'char') and key.char:
                char = key.char
            elif key == keyboard.Key.space:
                char = " "
            elif key == keyboard.Key.enter:
                char = "\n"
            elif key == keyboard.Key.backspace:
                is_backspace = True
            
            if char is not None or is_backspace:
                now = time.time()
                can_merge = False
                
                # Só agrupamos se a última ação gravada for digitação e recente (menos de 3 segundos)
                if self.current_clicks:
                    last_action = self.current_clicks[-1]
                    if last_action.type == Action.TYPE_TYPE:
                        time_diff = now - last_action.timestamp
                        if time_diff < 3.0:
                            can_merge = True
                
                if can_merge:
                    last_action = self.current_clicks[-1]
                    if is_backspace:
                        if last_action.text:
                            last_action.text = last_action.text[:-1]
                    else:
                        last_action.text += char
                    # Atualiza o timestamp para manter a janela ativa
                    last_action.timestamp = now
                else:
                    if not is_backspace:
                        action = Action(
                            Action.TYPE_TYPE,
                            timestamp=now,
                            text=char
                        )
                        self.current_clicks.append(action)
                
                self.current_count_var.set(f"Contador: {len(self.current_clicks)} ações registradas")
        except Exception as e:
            print(f"Erro ao capturar tecla: {e}")

    def on_key_release(self, key):
        """Captura liberação de teclas (pode ser usado para outros fins)."""
        pass

    def stop_recording(self):
        if not self.recording:
            return

        self.recording = False
        self.stop_button.config(state=tk.DISABLED)
        self.start_button.config(state=tk.NORMAL)

        if self.listener is not None:
            self.listener.stop()
            self.listener = None
        
        if self.keyboard_listener is not None:
            self.keyboard_listener.stop()
            self.keyboard_listener = None

        if not self.current_clicks:
            self.status_var.set("Nenhuma ação foi registrada. Inicie novamente ou verifique se o mouse/teclado estava ativo.")
            return

        name = simpledialog.askstring(
            "Nome da gravação",
            "Digite um nome para esta gravação:",
            parent=self.root,
        )
        if not name:
            name = f"Gravação {len(self.recordings) + 1}"

        # Salva no banco de dados
        try:
            if self.db.add_recording(name, self.current_clicks.copy()):
                self.load_recordings()
                self.status_var.set(f"Gravação '{name}' salva com {len(self.current_clicks)} ações.")
            else:
                self.status_var.set(f"Erro ao salvar gravação '{name}'.")
        except ValueError as e:
            messagebox.showerror("Erro", str(e))
            self.status_var.set("Nome de gravação já existe. Tente outro nome.")

    def create_logo(self, parent):
        """Carrega logo visual a partir de arquivo PNG/JPEG."""
        try:
            # Carrega a imagem
            img_path = r"C:\Users\willi\Documents\danilo\wbo-tecnologia.png"
            img = Image.open(img_path)
            
            # Redimensiona para caber no frame
            img.thumbnail((60, 60), Image.Resampling.LANCZOS)
            self.photo_image = ImageTk.PhotoImage(img)
            
            # Cria label com imagem
            img_label = tk.Label(parent, image=self.photo_image, bg="white")
            img_label.pack(side=tk.LEFT, padx=10, pady=5)
            
        except Exception as e:
            # Se falhar, cria um placeholder
            tk.Label(parent, text="Logo", font=("Arial", 16, "bold"), bg="white", fg="#0097A7", width=8).pack(side=tk.LEFT, padx=10, pady=5)
            print(f"Aviso: Não foi possível carregar a imagem. {e}")

        # Adiciona título e descrição
        title_frame = tk.Frame(parent, bg="white")
        title_frame.pack(side=tk.LEFT, padx=5, pady=5)

        tk.Label(title_frame, text="Click Recorder", font=("Arial", 14, "bold"), bg="white", fg="#0097A7").pack(anchor=tk.W)
        tk.Label(title_frame, text="Sistema de Gravação de Clicks", font=("Arial", 9), bg="white", fg="#666").pack(anchor=tk.W)

    def update_recordings_listbox(self):
        self.recordings_listbox.delete(0, tk.END)
        for index, record in enumerate(self.recordings, start=1):
            clicks_count = len(json.loads(record['clicks_data'])) if isinstance(record['clicks_data'], str) else len(record.get('clicks', []))
            run_cnt = record.get('run_count', 0)
            if run_cnt is None:
                run_cnt = 0
            label = f"{index}. {record['name']} ({clicks_count} passos) [Executado: {run_cnt}x]"
            self.recordings_listbox.insert(tk.END, label)

    def load_recordings(self):
        """Carrega todas as gravações do banco de dados."""
        self.recordings = self.db.get_all_recordings()
        self.update_recordings_listbox()

    def new_automation(self):
        """Cria uma nova automação vazia e abre o editor unificado."""
        name = simpledialog.askstring(
            "Nova Automação",
            "Digite um nome para a nova automação:",
            parent=self.root,
        )
        if not name:
            return
            
        name = name.strip()
        if not name:
            return
            
        try:
            # Salva uma lista vazia de passos no banco
            if self.db.add_recording(name, []):
                self.load_recordings()
                self.status_var.set(f"Automação '{name}' criada do zero.")
                
                # Encontra o índice da automação recém-criada para abri-la no editor
                for idx, record in enumerate(self.recordings):
                    if record['name'] == name:
                        # Seleciona na listbox
                        self.recordings_listbox.selection_clear(0, tk.END)
                        self.recordings_listbox.selection_set(idx)
                        self.recordings_listbox.see(idx)
                        # Abre o editor unificado
                        self.edit_automation()
                        break
            else:
                self.status_var.set("Erro ao criar a automação.")
        except ValueError as e:
            messagebox.showerror("Erro", str(e))

    def edit_automation(self):
        """Abre o editor unificado de automação (renomear e editar passos)."""
        selected = self.recordings_listbox.curselection()
        if not selected:
            messagebox.showinfo("Editar Automação", "Selecione uma automação para editar.")
            return

        index = selected[0]
        record = self.recordings[index]
        actions_data = json.loads(record['clicks_data']) if isinstance(record['clicks_data'], str) else record.get('clicks', [])
        actions = [Action.from_dict(a) if isinstance(a, dict) else a for a in actions_data]

        edit_window = tk.Toplevel(self.root)
        edit_window.title(f"Editar Automação: {record['name']}")
        edit_window.geometry("700x620")
        edit_window.resizable(True, True)

        # Nome da Automação
        tk.Label(edit_window, text="Nome da Automação:", font=("Arial", 11, "bold")).pack(anchor=tk.W, padx=15, pady=(15, 3))
        name_var = tk.StringVar(value=record["name"])
        tk.Entry(edit_window, textvariable=name_var, width=60, font=("Arial", 10)).pack(padx=15, pady=(0, 10), anchor=tk.W)

        # Listbox de ações/passos
        tk.Label(edit_window, text="Passos da Automação:", font=("Arial", 11, "bold")).pack(anchor=tk.W, padx=15, pady=(10, 5))
        
        listbox_frame = tk.Frame(edit_window)
        listbox_frame.pack(padx=15, pady=5, fill=tk.BOTH, expand=True)

        scrollbar = tk.Scrollbar(listbox_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        actions_listbox = tk.Listbox(listbox_frame, yscrollcommand=scrollbar.set, font=("Arial", 10))
        actions_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=actions_listbox.yview)

        # Popula listbox
        for idx, action in enumerate(actions, start=1):
            actions_listbox.insert(tk.END, f"{idx}. {str(action)}")

        button_frame = tk.Frame(edit_window)
        button_frame.pack(padx=15, pady=10, fill=tk.X)

        def open_edit_dialog(action, is_new=False, idx=None, insert_idx=None):
            edit_action_window = tk.Toplevel(edit_window)
            if is_new:
                edit_action_window.title("Adicionar Novo Passo")
            else:
                edit_action_window.title(f"Editar Passo {idx + 1}")
            edit_action_window.geometry("480x420")
            edit_action_window.resizable(False, False)

            # Tipo de ação
            tk.Label(edit_action_window, text="Tipo de Passo:", font=("Arial", 10, "bold")).pack(anchor=tk.W, padx=15, pady=(15, 3))
            type_var = tk.StringVar(value=action.type)
            
            # Frame para os campos dinâmicos
            fields_frame = tk.Frame(edit_action_window)
            fields_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)

            # Variáveis para guardar os valores dos campos
            x_var = tk.StringVar(value=str(action.x))
            y_var = tk.StringVar(value=str(action.y))
            button_var = tk.StringVar(value=action.button)
            text_var = tk.StringVar(value=action.text)
            duration_var = tk.StringVar(value=str(action.duration))
            file_path_var = tk.StringVar(value=action.file_path)
            read_mode_var = tk.StringVar(value=action.read_mode)
            file_selector_var = tk.StringVar(value=action.file_selector)
            dx_var = tk.StringVar(value=str(action.dx))
            dy_var = tk.StringVar(value=str(action.dy))
            ai_task_var = tk.StringVar(value=action.ai_task)
            ai_mode_var = tk.StringVar(value=action.ai_mode)
            file_path_2_var = tk.StringVar(value=action.file_path_2)
            file_path_3_var = tk.StringVar(value=action.file_path_3)
            output_path_var = tk.StringVar(value=action.output_path)

            def browse_file(entry_widget, target_var=None):
                from tkinter import filedialog
                file_path = filedialog.askopenfilename(
                    title="Selecione o arquivo para leitura",
                    filetypes=[("Todos os arquivos", "*.*"), ("Text Files", "*.txt"), ("JSON Files", "*.json"), ("CSV Files", "*.csv")]
                )
                if file_path:
                    entry_widget.delete(0, tk.END)
                    entry_widget.insert(0, file_path)
                    if target_var is not None:
                        target_var.set(file_path)
                    else:
                        file_path_var.set(file_path)

            def draw_fields(*args):
                for widget in fields_frame.winfo_children():
                    widget.destroy()

                selected_type = type_var.get()
                
                # Ajuste de tamanho dinâmico da janela
                if selected_type == Action.TYPE_AI:
                    if ai_mode_var.get() == 'excel_merge':
                        edit_action_window.geometry("500x680")
                    else:
                        edit_action_window.geometry("500x560")
                else:
                    edit_action_window.geometry("480x420")
                
                if selected_type == Action.TYPE_CLICK:
                    tk.Label(fields_frame, text="Coordenada X:", font=("Arial", 10)).pack(anchor=tk.W, pady=(5, 2))
                    tk.Entry(fields_frame, textvariable=x_var, width=50).pack(anchor=tk.W, pady=(0, 5))

                    tk.Label(fields_frame, text="Coordenada Y:", font=("Arial", 10)).pack(anchor=tk.W, pady=(5, 2))
                    tk.Entry(fields_frame, textvariable=y_var, width=50).pack(anchor=tk.W, pady=(0, 5))

                    tk.Label(fields_frame, text="Botão do Mouse:", font=("Arial", 10)).pack(anchor=tk.W, pady=(5, 2))
                    tk.OptionMenu(fields_frame, button_var, 'Button.left', 'Button.right', 'Button.middle').pack(anchor=tk.W, pady=(0, 5))

                elif selected_type == Action.TYPE_TYPE:
                    tk.Label(fields_frame, text="Texto para digitar:", font=("Arial", 10)).pack(anchor=tk.W, pady=(5, 2))
                    entry = tk.Entry(fields_frame, textvariable=text_var, width=50)
                    entry.pack(anchor=tk.W, pady=(0, 5))

                elif selected_type == Action.TYPE_WAIT:
                    tk.Label(fields_frame, text="Duração (segundos):", font=("Arial", 10)).pack(anchor=tk.W, pady=(5, 2))
                    tk.Entry(fields_frame, textvariable=duration_var, width=50).pack(anchor=tk.W, pady=(0, 5))

                elif selected_type == Action.TYPE_READ_FILE:
                    tk.Label(fields_frame, text="Caminho do Arquivo:", font=("Arial", 10)).pack(anchor=tk.W, pady=(5, 2))
                    path_frame = tk.Frame(fields_frame)
                    path_frame.pack(fill=tk.X, anchor=tk.W, pady=(0, 5))
                    path_entry = tk.Entry(path_frame, textvariable=file_path_var, width=40)
                    path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
                    tk.Button(path_frame, text="Procurar...", command=lambda: browse_file(path_entry)).pack(side=tk.LEFT, padx=(5, 0))

                    tk.Label(fields_frame, text="Modo de Leitura:", font=("Arial", 10)).pack(anchor=tk.W, pady=(5, 2))
                    mode_options = {
                        'complete': 'Conteúdo Completo',
                        'line': 'Linha Específica',
                        'json': 'Chave JSON',
                        'csv': 'Linha/Coluna CSV'
                    }
                    display_mode_var = tk.StringVar()
                    reverse_mode_map = {v: k for k, v in mode_options.items()}
                    
                    display_mode_var.set(mode_options.get(read_mode_var.get(), 'Conteúdo Completo'))
                    
                    def on_mode_change(*a):
                        read_mode_var.set(reverse_mode_map[display_mode_var.get()])
                        update_helper_label()

                    display_mode_var.trace_add("write", on_mode_change)
                    tk.OptionMenu(fields_frame, display_mode_var, *mode_options.values()).pack(anchor=tk.W, pady=(0, 5))

                    tk.Label(fields_frame, text="Seletor/Índice:", font=("Arial", 10)).pack(anchor=tk.W, pady=(5, 2))
                    tk.Entry(fields_frame, textvariable=file_selector_var, width=50).pack(anchor=tk.W, pady=(0, 2))
                    
                    helper_lbl = tk.Label(fields_frame, text="", fg="gray", font=("Arial", 8))
                    helper_lbl.pack(anchor=tk.W)

                    def update_helper_label():
                        m = read_mode_var.get()
                        if m == 'complete':
                            helper_lbl.config(text="Não aplicável para conteúdo completo.")
                        elif m == 'line':
                            helper_lbl.config(text="Ex: 1 para primeira linha, 2 para segunda, etc.")
                        elif m == 'json':
                            helper_lbl.config(text="Ex: chave ou chave_aninhada.campo (ex: usuario.nome)")
                        elif m == 'csv':
                            helper_lbl.config(text="Ex: linha,coluna (ex: 1,2) ou nome da coluna")

                    update_helper_label()

                elif selected_type == Action.TYPE_SCROLL:
                    tk.Label(fields_frame, text="Deslocamento Horizontal (dx):", font=("Arial", 10)).pack(anchor=tk.W, pady=(5, 2))
                    tk.Entry(fields_frame, textvariable=dx_var, width=50).pack(anchor=tk.W, pady=(0, 5))

                    tk.Label(fields_frame, text="Deslocamento Vertical (dy):", font=("Arial", 10)).pack(anchor=tk.W, pady=(5, 2))
                    tk.Entry(fields_frame, textvariable=dy_var, width=50).pack(anchor=tk.W, pady=(0, 5))
                    
                    tk.Label(fields_frame, text="Dica: dy positivo rola para cima, dy negativo rola para baixo.\ndx positivo rola para a direita, dx negativo rola para a esquerda.", fg="gray", font=("Arial", 8), justify=tk.LEFT).pack(anchor=tk.W, pady=(5, 0))

                elif selected_type == Action.TYPE_AI:
                    tk.Label(fields_frame, text="Modo de IA:", font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=(5, 2))
                    mode_options = {
                        'ocr': 'Visão/OCR (Desktop)',
                        'selenium': 'Selenium Web (Navegador)',
                        'excel_merge': 'Mesclar Arquivos para Excel'
                    }
                    display_ai_mode = tk.StringVar()
                    reverse_ai_mode_map = {v: k for k, v in mode_options.items()}
                    display_ai_mode.set(mode_options.get(ai_mode_var.get(), 'Visão/OCR (Desktop)'))
                    
                    def on_ai_mode_change(*a):
                        ai_mode_var.set(reverse_ai_mode_map[display_ai_mode.get()])
                        draw_fields()

                    display_ai_mode.trace_add("write", on_ai_mode_change)
                    tk.OptionMenu(fields_frame, display_ai_mode, *mode_options.values()).pack(anchor=tk.W, pady=(0, 5))

                    # Perfis de IA (Skills)
                    profile_frame = tk.Frame(fields_frame)
                    profile_frame.pack(fill=tk.X, anchor=tk.W, pady=(5, 5))
                    
                    tk.Label(profile_frame, text="Perfil de IA (Skill):", font=("Arial", 9, "bold")).pack(side=tk.LEFT, anchor=tk.W)
                    
                    skills = self.db.get_all_skills()
                    skill_map = {s["name"]: s["prompt"] for s in skills}
                    skill_names = sorted(list(skill_map.keys()))
                    
                    selected_skill_var = tk.StringVar(value="Selecione um perfil...")
                    options_list = ["Selecione um perfil..."] + skill_names
                    
                    def on_skill_selected(*a):
                        sel = selected_skill_var.get()
                        if sel in skill_map:
                            task_text.delete("1.0", tk.END)
                            task_text.insert(tk.END, skill_map[sel])
                            ai_task_var.set(skill_map[sel])
                    
                    selected_skill_var.trace_add("write", on_skill_selected)
                    
                    profile_menu = tk.OptionMenu(profile_frame, selected_skill_var, *options_list)
                    profile_menu.pack(side=tk.LEFT, padx=5)
                    
                    # Botões de perfil
                    btn_profile_frame = tk.Frame(fields_frame)
                    btn_profile_frame.pack(fill=tk.X, anchor=tk.W, pady=(0, 5))
                    
                    def save_profile():
                        current_prompt = task_text.get("1.0", tk.END).strip()
                        if not current_prompt:
                            messagebox.showwarning("Aviso", "Por favor, insira alguma instrução para salvar como perfil.", parent=edit_action_window)
                            return
                        
                        profile_name = simpledialog.askstring(
                            "Salvar Perfil",
                            "Digite o nome para este perfil de IA (Skill):",
                            parent=edit_action_window
                        )
                        if profile_name:
                            profile_name = profile_name.strip()
                            if profile_name:
                                if self.db.add_skill(profile_name, current_prompt):
                                    messagebox.showinfo("Sucesso", f"Perfil '{profile_name}' salvo com sucesso!", parent=edit_action_window)
                                    draw_fields()
                                else:
                                    messagebox.showerror("Erro", "Falha ao salvar perfil no banco de dados.", parent=edit_action_window)
                                    
                    def delete_profile():
                        sel = selected_skill_var.get()
                        if sel == "Selecione um perfil..." or sel not in skill_map:
                            messagebox.showwarning("Aviso", "Selecione um perfil válido para excluir.", parent=edit_action_window)
                            return
                        
                        confirm = messagebox.askyesno(
                            "Confirmar Exclusão",
                            f"Tem certeza que deseja excluir o perfil '{sel}'?",
                            parent=edit_action_window
                        )
                        if confirm:
                            if self.db.delete_skill(sel):
                                messagebox.showinfo("Sucesso", f"Perfil '{sel}' excluído com sucesso!", parent=edit_action_window)
                                draw_fields()
                            else:
                                messagebox.showerror("Erro", f"Erro ao excluir perfil '{sel}'.", parent=edit_action_window)
                                
                    tk.Button(btn_profile_frame, text="Salvar como Perfil", command=save_profile, bg="#4caf50", fg="white", font=("Arial", 8, "bold")).pack(side=tk.LEFT, padx=(0, 5))
                    tk.Button(btn_profile_frame, text="Excluir Perfil Selecionado", command=delete_profile, bg="#f44336", fg="white", font=("Arial", 8, "bold")).pack(side=tk.LEFT)

                    tk.Label(fields_frame, text="Tarefa/Skill da IA (Instruções):", font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=(5, 2))
                    task_text = tk.Text(fields_frame, width=55, height=3, font=("Arial", 9))
                    task_text.pack(anchor=tk.W, pady=(0, 5))
                    task_text.insert(tk.END, ai_task_var.get())
                    
                    def on_task_key(event):
                        ai_task_var.set(task_text.get("1.0", tk.END).strip())
                    task_text.bind("<KeyRelease>", on_task_key)

                    current_mode = ai_mode_var.get()

                    if current_mode == 'excel_merge':
                        tk.Label(fields_frame, text="Arquivo 1 (Origem):", font=("Arial", 9)).pack(anchor=tk.W, pady=(2, 1))
                        f1_frame = tk.Frame(fields_frame)
                        f1_frame.pack(fill=tk.X, anchor=tk.W, pady=(0, 3))
                        f1_entry = tk.Entry(f1_frame, textvariable=file_path_var, width=40)
                        f1_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
                        tk.Button(f1_frame, text="Procurar...", command=lambda: browse_file(f1_entry, file_path_var)).pack(side=tk.LEFT, padx=(5, 0))

                        tk.Label(fields_frame, text="Arquivo 2 (Origem):", font=("Arial", 9)).pack(anchor=tk.W, pady=(2, 1))
                        f2_frame = tk.Frame(fields_frame)
                        f2_frame.pack(fill=tk.X, anchor=tk.W, pady=(0, 3))
                        f2_entry = tk.Entry(f2_frame, textvariable=file_path_2_var, width=40)
                        f2_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
                        tk.Button(f2_frame, text="Procurar...", command=lambda: browse_file(f2_entry, file_path_2_var)).pack(side=tk.LEFT, padx=(5, 0))

                        tk.Label(fields_frame, text="Arquivo 3 (Origem):", font=("Arial", 9)).pack(anchor=tk.W, pady=(2, 1))
                        f3_frame = tk.Frame(fields_frame)
                        f3_frame.pack(fill=tk.X, anchor=tk.W, pady=(0, 3))
                        f3_entry = tk.Entry(f3_frame, textvariable=file_path_3_var, width=40)
                        f3_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
                        tk.Button(f3_frame, text="Procurar...", command=lambda: browse_file(f3_entry, file_path_3_var)).pack(side=tk.LEFT, padx=(5, 0))

                        tk.Label(fields_frame, text="Caminho do Excel Gerado (Saída):", font=("Arial", 9)).pack(anchor=tk.W, pady=(2, 1))
                        out_frame = tk.Frame(fields_frame)
                        out_frame.pack(fill=tk.X, anchor=tk.W, pady=(0, 3))
                        out_entry = tk.Entry(out_frame, textvariable=output_path_var, width=40)
                        out_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
                        
                        def browse_save_file():
                            from tkinter import filedialog
                            out_p = filedialog.asksaveasfilename(
                                title="Salvar Excel formatado",
                                filetypes=[("Excel Files", "*.xlsx")],
                                defaultextension=".xlsx"
                            )
                            if out_p:
                                out_entry.delete(0, tk.END)
                                out_entry.insert(0, out_p)
                                output_path_var.set(out_p)

                        tk.Button(out_frame, text="Salvar em...", command=browse_save_file).pack(side=tk.LEFT, padx=(5, 0))
                    else:
                        tk.Label(fields_frame, text="Arquivo de Dados Origem (opcional):", font=("Arial", 9)).pack(anchor=tk.W, pady=(5, 2))
                        ds_frame = tk.Frame(fields_frame)
                        ds_frame.pack(fill=tk.X, anchor=tk.W, pady=(0, 5))
                        ds_entry = tk.Entry(ds_frame, textvariable=file_path_var, width=40)
                        ds_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
                        tk.Button(ds_frame, text="Procurar...", command=lambda: browse_file(ds_entry, file_path_var)).pack(side=tk.LEFT, padx=(5, 0))

            # Para gerenciar listeners temporários de captura
            temp_listeners = []
            
            def cleanup_listeners(event=None):
                if event is not None and event.widget != edit_action_window:
                    return
                for listener in temp_listeners:
                    try:
                        listener.stop()
                    except Exception:
                        pass
                temp_listeners.clear()

            edit_action_window.bind("<Destroy>", cleanup_listeners)

            def start_temp_capture():
                capture_btn.config(state=tk.DISABLED)
                
                def is_inside_window(window, x, y):
                    try:
                        if not window.winfo_exists():
                            return False
                        win_x = window.winfo_rootx()
                        win_y = window.winfo_rooty()
                        win_w = window.winfo_width()
                        win_h = window.winfo_height()
                        return win_x <= x <= win_x + win_w and win_y <= y <= win_y + win_h
                    except Exception:
                        return False

                def update_btn_text(text):
                    if edit_action_window.winfo_exists():
                        capture_btn.config(text=text)

                def run_countdown():
                    try:
                        for count in [3, 2, 1]:
                            update_btn_text(f"Capturando em {count}...")
                            time.sleep(1)
                        
                        update_btn_text("Aguardando ação...")
                        
                        def on_temp_click(x, y, button, pressed):
                            if pressed:
                                if (is_inside_window(self.root, x, y) or 
                                    is_inside_window(edit_window, x, y) or 
                                    is_inside_window(edit_action_window, x, y)):
                                    return
                                
                                cleanup_listeners()
                                
                                def apply_click():
                                    type_var.set(Action.TYPE_CLICK)
                                    x_var.set(str(x))
                                    y_var.set(str(y))
                                    button_var.set(str(button))
                                    draw_fields()
                                    capture_btn.config(state=tk.NORMAL, text="Capturar Passo")
                                
                                edit_action_window.after(0, apply_click)

                        def on_temp_scroll(x, y, dx, dy):
                            if (is_inside_window(self.root, x, y) or 
                                is_inside_window(edit_window, x, y) or 
                                is_inside_window(edit_action_window, x, y)):
                                return
                            
                            cleanup_listeners()
                            
                            def apply_scroll():
                                type_var.set(Action.TYPE_SCROLL)
                                dx_var.set(str(dx))
                                dy_var.set(str(dy))
                                draw_fields()
                                capture_btn.config(state=tk.NORMAL, text="Capturar Passo")
                            
                            edit_action_window.after(0, apply_scroll)

                        def on_temp_key_press(key):
                            char = None
                            if hasattr(key, 'char') and key.char:
                                char = key.char
                            elif key == keyboard.Key.space:
                                char = " "
                            elif key == keyboard.Key.enter:
                                char = "\n"
                            
                            if char is None:
                                return
                                
                            cleanup_listeners()
                            
                            def apply_type():
                                type_var.set(Action.TYPE_TYPE)
                                text_var.set(char)
                                draw_fields()
                                capture_btn.config(state=tk.NORMAL, text="Capturar Passo")
                            
                            edit_action_window.after(0, apply_type)

                        m_listener = mouse.Listener(on_click=on_temp_click, on_scroll=on_temp_scroll)
                        k_listener = keyboard.Listener(on_press=on_temp_key_press)
                        
                        temp_listeners.extend([m_listener, k_listener])
                        m_listener.start()
                        k_listener.start()
                        
                    except Exception as e:
                        print(f"Erro na captura temporária: {e}")
                        def reset_btn():
                            capture_btn.config(state=tk.NORMAL, text="Capturar Passo")
                        edit_action_window.after(0, reset_btn)

                threading.Thread(target=run_countdown, daemon=True).start()

            type_frame = tk.Frame(edit_action_window)
            type_frame.pack(anchor=tk.W, padx=15, pady=(0, 10))

            type_menu = tk.OptionMenu(type_frame, type_var, Action.TYPE_CLICK, Action.TYPE_TYPE, Action.TYPE_WAIT, Action.TYPE_READ_FILE, Action.TYPE_SCROLL, Action.TYPE_AI)
            type_menu.pack(side=tk.LEFT)
            
            capture_btn = tk.Button(type_frame, text="Capturar Passo", command=start_temp_capture, bg="#9c27b0", fg="white", font=("Arial", 9, "bold"))
            capture_btn.pack(side=tk.LEFT, padx=(10, 0))
            
            type_var.trace_add("write", draw_fields)
            draw_fields()

            def save_action_changes():
                action_type = type_var.get()
                
                # Validações
                if action_type == Action.TYPE_CLICK:
                    try:
                        val_x = int(x_var.get())
                    except ValueError:
                        messagebox.showwarning("Validação", "A coordenada X deve ser um número inteiro.", parent=edit_action_window)
                        return
                    try:
                        val_y = int(y_var.get())
                    except ValueError:
                        messagebox.showwarning("Validação", "A coordenada Y deve ser um número inteiro.", parent=edit_action_window)
                        return
                        
                    action.type = action_type
                    action.x = val_x
                    action.y = val_y
                    action.button = button_var.get()
                    
                elif action_type == Action.TYPE_TYPE:
                    action.type = action_type
                    action.text = text_var.get()
                    
                elif action_type == Action.TYPE_WAIT:
                    try:
                        val_dur = float(duration_var.get())
                        if val_dur < 0:
                            raise ValueError()
                    except ValueError:
                        messagebox.showwarning("Validação", "A duração deve ser um número maior ou igual a 0.", parent=edit_action_window)
                        return
                        
                    action.type = action_type
                    action.duration = val_dur
                    
                elif action_type == Action.TYPE_READ_FILE:
                    path_val = file_path_var.get().strip()
                    if not path_val:
                        messagebox.showwarning("Validação", "Por favor, selecione um arquivo de leitura.", parent=edit_action_window)
                        return
                    
                    mode_val = read_mode_var.get()
                    selector_val = file_selector_var.get().strip()
                    
                    if mode_val == 'line':
                        try:
                            val_ln = int(selector_val)
                            if val_ln <= 0:
                                raise ValueError()
                        except ValueError:
                            messagebox.showwarning("Validação", "No modo 'Linha Específica', o seletor deve ser um número de linha maior que 0.", parent=edit_action_window)
                            return
                            
                    action.type = action_type
                    action.file_path = path_val
                    action.read_mode = mode_val
                    action.file_selector = selector_val
                
                elif action_type == Action.TYPE_SCROLL:
                    try:
                        val_dx = int(dx_var.get())
                    except ValueError:
                        messagebox.showwarning("Validação", "O deslocamento dx deve ser um número inteiro.", parent=edit_action_window)
                        return
                    try:
                        val_dy = int(dy_var.get())
                    except ValueError:
                        messagebox.showwarning("Validação", "O deslocamento dy deve ser um número inteiro.", parent=edit_action_window)
                        return
                        
                    action.type = action_type
                    action.dx = val_dx
                    action.dy = val_dy

                elif action_type == Action.TYPE_AI:
                    task_val = ai_task_var.get().strip()
                    if not task_val:
                        messagebox.showwarning("Validação", "Por favor, digite a tarefa ou skill da IA.", parent=edit_action_window)
                        return
                    
                    mode_val = ai_mode_var.get()
                    if mode_val == 'excel_merge':
                        if not file_path_var.get().strip() and not file_path_2_var.get().strip() and not file_path_3_var.get().strip():
                            messagebox.showwarning("Validação", "Por favor, selecione ao menos um arquivo de origem para mesclar.", parent=edit_action_window)
                            return
                        if not output_path_var.get().strip():
                            messagebox.showwarning("Validação", "Por favor, defina o caminho do arquivo Excel de saída.", parent=edit_action_window)
                            return
                            
                    action.type = action_type
                    action.ai_task = task_val
                    action.ai_mode = mode_val
                    action.file_path = file_path_var.get().strip()
                    action.file_path_2 = file_path_2_var.get().strip()
                    action.file_path_3 = file_path_3_var.get().strip()
                    action.output_path = output_path_var.get().strip()
                
                try:
                    if is_new:
                        if insert_idx is not None:
                            actions.insert(insert_idx, action)
                        else:
                            actions.append(action)
                    
                    # Atualiza os índices visuais de todos na listbox
                    actions_listbox.delete(0, tk.END)
                    for i, act in enumerate(actions, start=1):
                        actions_listbox.insert(tk.END, f"{i}. {str(act)}")
                    
                    # Seleciona e foca na nova ação inserida ou na editada
                    actions_listbox.selection_clear(0, tk.END)
                    select_idx = insert_idx if is_new else idx
                    if select_idx is not None and select_idx < len(actions):
                        actions_listbox.selection_set(select_idx)
                        actions_listbox.see(select_idx)
                        
                    edit_action_window.destroy()
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao salvar passo: {e}")

            tk.Button(edit_action_window, text="Salvar Passo", command=save_action_changes, width=20, bg="#4caf50", fg="white", font=("Arial", 10, "bold")).pack(pady=15)

        def edit_action():
            selection = actions_listbox.curselection()
            if not selection:
                messagebox.showinfo("Editar Passo", "Selecione um passo para editar.")
                return
            action_idx = selection[0]
            action = actions[action_idx]
            open_edit_dialog(action, is_new=False, idx=action_idx)

        def delete_action():
            selection = actions_listbox.curselection()
            if not selection:
                messagebox.showinfo("Deletar Passo", "Selecione um passo para deletar.")
                return
            action_idx = selection[0]
            del actions[action_idx]
            
            actions_listbox.delete(0, tk.END)
            for i, act in enumerate(actions, start=1):
                actions_listbox.insert(tk.END, f"{i}. {str(act)}")

        def duplicate_action():
            selection = actions_listbox.curselection()
            if not selection:
                messagebox.showinfo("Duplicar Passo", "Selecione um passo para duplicar.")
                return
            action_idx = selection[0]
            action_to_dup = actions[action_idx]
            new_action = Action.from_dict(action_to_dup.to_dict())
            
            actions.insert(action_idx + 1, new_action)
            
            actions_listbox.delete(0, tk.END)
            for i, act in enumerate(actions, start=1):
                actions_listbox.insert(tk.END, f"{i}. {str(act)}")
                
            actions_listbox.selection_clear(0, tk.END)
            actions_listbox.selection_set(action_idx + 1)
            actions_listbox.see(action_idx + 1)

        def move_up():
            selection = actions_listbox.curselection()
            if not selection:
                messagebox.showinfo("Mover Passo", "Selecione um passo para mover.")
                return
            idx = selection[0]
            if idx > 0:
                actions[idx], actions[idx - 1] = actions[idx - 1], actions[idx]
                
                actions_listbox.delete(0, tk.END)
                for i, act in enumerate(actions, start=1):
                    actions_listbox.insert(tk.END, f"{i}. {str(act)}")
                    
                actions_listbox.selection_clear(0, tk.END)
                actions_listbox.selection_set(idx - 1)
                actions_listbox.see(idx - 1)

        def move_down():
            selection = actions_listbox.curselection()
            if not selection:
                messagebox.showinfo("Mover Passo", "Selecione um passo para mover.")
                return
            idx = selection[0]
            if idx < len(actions) - 1:
                actions[idx], actions[idx + 1] = actions[idx + 1], actions[idx]
                
                actions_listbox.delete(0, tk.END)
                for i, act in enumerate(actions, start=1):
                    actions_listbox.insert(tk.END, f"{i}. {str(act)}")
                    
                actions_listbox.selection_clear(0, tk.END)
                actions_listbox.selection_set(idx + 1)
                actions_listbox.see(idx + 1)

        def get_insert_index():
            selection = actions_listbox.curselection()
            if selection:
                return selection[0] + 1
            return len(actions)

        def add_click():
            new_action = Action(Action.TYPE_CLICK, x=0, y=0)
            open_edit_dialog(new_action, is_new=True, insert_idx=get_insert_index())

        def add_type():
            new_action = Action(Action.TYPE_TYPE, text="")
            open_edit_dialog(new_action, is_new=True, insert_idx=get_insert_index())

        def add_wait():
            new_action = Action(Action.TYPE_WAIT, duration=1.0)
            open_edit_dialog(new_action, is_new=True, insert_idx=get_insert_index())

        def add_read_file():
            new_action = Action(Action.TYPE_READ_FILE, file_path="", read_mode="complete", file_selector="")
            open_edit_dialog(new_action, is_new=True, insert_idx=get_insert_index())

        def add_scroll():
            new_action = Action(Action.TYPE_SCROLL, dx=0, dy=0)
            open_edit_dialog(new_action, is_new=True, insert_idx=get_insert_index())

        def add_ai_task():
            new_action = Action(Action.TYPE_AI, ai_task="", ai_mode="ocr")
            open_edit_dialog(new_action, is_new=True, insert_idx=get_insert_index())

        # Fileira 1: Operações no passo selecionado
        selected_ops_frame = tk.Frame(button_frame)
        selected_ops_frame.pack(fill=tk.X, pady=(0, 5))
        
        tk.Button(selected_ops_frame, text="Editar Selecionado", command=edit_action, width=16, bg="#2196f3", fg="white", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(selected_ops_frame, text="Deletar Selecionado", command=delete_action, width=16, bg="#f44336", fg="white", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(selected_ops_frame, text="Duplicar", command=duplicate_action, width=10, bg="#ff9800", fg="white", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(selected_ops_frame, text="↑ Subir", command=move_up, width=8, bg="#757575", fg="white", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(selected_ops_frame, text="↓ Descer", command=move_down, width=8, bg="#757575", fg="white", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)
        
        # Fileira 2: Adicionar novos passos
        add_ops_frame = tk.Frame(button_frame)
        add_ops_frame.pack(fill=tk.X)
        
        tk.Button(add_ops_frame, text="+ Clique", command=add_click, width=10, bg="#4caf50", fg="white", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(add_ops_frame, text="+ Digitar", command=add_type, width=10, bg="#00bcd4", fg="white", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(add_ops_frame, text="+ Espera", command=add_wait, width=10, bg="#ff9800", fg="white", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(add_ops_frame, text="+ Ler Arquivo", command=add_read_file, width=12, bg="#673ab7", fg="white", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(add_ops_frame, text="+ Rolar", command=add_scroll, width=10, bg="#9c27b0", fg="white", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(add_ops_frame, text="+ Passo IA", command=add_ai_task, width=12, bg="#e91e63", fg="white", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)

        def save_all_changes():
            """Salva todas as alterações (nome e passos) no banco de dados."""
            new_name = name_var.get().strip() or record["name"]
            try:
                name_changed = (new_name != record["name"])
                clicks_list = [a.to_dict() for a in actions]
                
                if self.db.update_recording(record["name"], new_name=new_name if name_changed else None, clicks=clicks_list):
                    self.load_recordings()
                    self.status_var.set(f"Automação '{new_name}' atualizada com sucesso.")
                    edit_window.destroy()
                else:
                    messagebox.showerror("Erro", "Falha ao salvar as alterações no banco de dados.")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar alterações: {e}")

        # Botão salvar geral no rodapé da janela principal de edição
        save_frame = tk.Frame(edit_window)
        save_frame.pack(padx=15, pady=15, fill=tk.X)
        tk.Button(save_frame, text="Salvar Todas as Alterações", command=save_all_changes, width=35, bg="#4caf50", fg="white", font=("Arial", 11, "bold")).pack()

    def delete_recording(self):
        selected = self.recordings_listbox.curselection()
        if not selected:
            messagebox.showinfo("Excluir automação", "Selecione uma automação para excluir.")
            return

        index = selected[0]
        record = self.recordings[index]
        confirm = messagebox.askyesno("Excluir automação", f"Remover '{record['name']}'?")
        if confirm:
            if self.db.delete_recording(record["name"]):
                self.load_recordings()
                self.status_var.set(f"Automação '{record['name']}' removida.")
            else:
                messagebox.showerror("Erro", "Falha ao deletar automação.")

    def stop_execution(self):
        """Para a execução automática ativa."""
        if self.playing:
            self.playing = False
            self.status_var.set("Solicitando parada da automação...")

    def execute_recording(self):
        if self.playing:
            messagebox.showwarning("Execução em andamento", "Uma gravação já está sendo executada.")
            return

        selected = self.recordings_listbox.curselection()
        if not selected:
            messagebox.showinfo("Executar gravação", "Selecione uma gravação para executar.")
            return

        index = selected[0]
        record = self.recordings[index]
        clicks = json.loads(record['clicks_data']) if isinstance(record['clicks_data'], str) else record.get('clicks', [])
        
        confirm = messagebox.askyesno("Executar gravação", f"Reproduzir '{record['name']}'?\n\nCliques: {len(clicks)}")
        if confirm:
            self.playing = True
            self.execute_button.config(state=tk.DISABLED)
            if hasattr(self, 'stop_execution_button'):
                self.stop_execution_button.config(state=tk.NORMAL)
            
            # Incrementa o run_count no banco e recarrega a interface
            self.db.increment_run_count(record['name'])
            self.load_recordings()
            self.recordings_listbox.selection_set(index)
            
            threading.Thread(target=self.playback_clicks, args=(clicks,), daemon=True).start()

    def playback_clicks(self, clicks):
        if not clicks:
            self.status_var.set("Nenhuma ação para reproduzir.")
            self.playing = False
            self.execute_button.config(state=tk.NORMAL)
            if hasattr(self, 'stop_execution_button'):
                self.stop_execution_button.config(state=tk.DISABLED)
            return

        try:
            # Converte dicts para Action objects se necessário
            actions = []
            for item in clicks:
                if isinstance(item, dict):
                    actions.append(Action.from_dict(item))
                else:
                    actions.append(item)

            controller = Controller()
            kb_controller = keyboard.Controller()
            
            self.status_var.set(f"Reproduzindo: 0/{len(actions)} ações...")
            time.sleep(1)

            last_timestamp = actions[0].timestamp if actions else time.time()
            
            for idx, action in enumerate(actions):
                if not self.playing:
                    break

                # Calcula delay relativo com limites de configuração
                if idx > 0:
                    delay = action.timestamp - last_timestamp
                    if delay > 0:
                        try:
                            max_wait = float(self.max_wait_var.get())
                        except Exception:
                            max_wait = 10.0
                        
                        if 0.35 <= delay < 0.5:
                            delay = 0.5
                        elif delay > max_wait:
                            delay = max_wait
                        # Dorme em pequenos passos para garantir resposta rápida ao botão Parar
                        steps = int(delay / 0.1)
                        for _ in range(steps):
                            if not self.playing:
                                break
                            time.sleep(0.1)
                        remainder = delay % 0.1
                        if self.playing and remainder > 0:
                            time.sleep(remainder)

                last_timestamp = action.timestamp

                if not self.playing:
                    break

                # Executa ação
                if action.type == Action.TYPE_CLICK:
                    button_str = action.button.lower()
                    if "right" in button_str:
                        button = Button.right
                    elif "middle" in button_str:
                        button = Button.middle
                    else:
                        button = Button.left

                    controller.position = (action.x, action.y)
                    time.sleep(0.05)
                    controller.press(button)
                    time.sleep(0.05)
                    controller.release(button)
                
                elif action.type == Action.TYPE_TYPE:
                    kb_controller.type(action.text)
                
                elif action.type == Action.TYPE_WAIT:
                    # Dorme em pequenos passos para responder rápido ao botão Parar
                    steps = int(action.duration / 0.1)
                    for _ in range(steps):
                        if not self.playing:
                            break
                        time.sleep(0.1)
                    remainder = action.duration % 0.1
                    if self.playing and remainder > 0:
                        time.sleep(remainder)
                
                elif action.type == Action.TYPE_READ_FILE:
                    try:
                        text_to_type = get_file_content_to_type(
                            action.file_path,
                            action.read_mode,
                            action.file_selector
                        )
                        kb_controller.type(text_to_type)
                    except Exception as err:
                        self.status_var.set(f"Erro ao ler arquivo: {str(err)}")
                        raise err

                elif action.type == Action.TYPE_SCROLL:
                    controller.scroll(action.dx, action.dy)

                elif action.type == Action.TYPE_AI:
                    self.status_var.set(f"Reproduzindo: {idx + 1}/{len(actions)} executando IA...")
                    self.run_autonomous_agent(
                        action.ai_task,
                        ai_mode=action.ai_mode,
                        file_path=action.file_path,
                        file_path_2=action.file_path_2,
                        file_path_3=action.file_path_3,
                        output_path=action.output_path
                    )

                self.status_var.set(f"Reproduzindo: {idx + 1}/{len(actions)} ações...")

            if self.playing:
                self.status_var.set(f"Automação reproduzida com sucesso: {len(actions)} ações executadas.")
            else:
                self.status_var.set("Execução interrompida pelo usuário.")

        except Exception as e:
            self.status_var.set(f"Erro ao reproduzir: {str(e)}")
            print(f"Erro detalhado: {e}")
        finally:
            self.playing = False
            self.execute_button.config(state=tk.NORMAL)
            if hasattr(self, 'stop_execution_button'):
                self.stop_execution_button.config(state=tk.DISABLED)

    @staticmethod
    def format_clicks(clicks):
        """Formata ações para exibição legível."""
        lines = []
        if not clicks:
            return "Nenhuma ação registrada."
        
        for idx, item in enumerate(clicks, start=1):
            if isinstance(item, dict):
                action = Action.from_dict(item)
            else:
                action = item
            lines.append(f"{idx}. {str(action)}")
        return "\n".join(lines)

    def append_chat_message(self, sender, text):
        self.chat_history.config(state=tk.NORMAL)
        self.chat_history.insert(tk.END, f"[{sender}]: {text}\n\n")
        self.chat_history.see(tk.END)
        self.chat_history.config(state=tk.DISABLED)
        # Salva no histórico local da conversa
        role = "user" if sender == "Você" else "assistant"
        if sender != "Sistema":
            self.chat_history_list.append({"role": role, "content": text})

    def send_user_message(self):
        msg = self.chat_input.get().strip()
        if not msg:
            return
        
        self.chat_input.delete(0, tk.END)
        self.append_chat_message("Você", msg)
        
        # Se o agente de IA estiver rodando e aguardando resposta
        if self.agent_active and not self.user_message_event.is_set():
            self.last_user_message = msg
            self.user_message_event.set()
        else:
            # Caso contrário, inicia uma nova tarefa autônoma livre
            if not self.playing:
                self.playing = True
                self.execute_button.config(state=tk.DISABLED)
                if hasattr(self, 'stop_execution_button'):
                    self.stop_execution_button.config(state=tk.NORMAL)
                
                # Inicia a thread do agente autônomo
                threading.Thread(target=self.run_autonomous_agent, args=(msg,), daemon=True).start()
            else:
                self.append_chat_message("Sistema", "Uma automação ou tarefa de agente já está rodando. Aguarde ou clique em Parar.")

    def run_autonomous_agent(self, task_description, ai_mode="ocr", file_path="", file_path_2="", file_path_3="", output_path=""):
        api_key = self.db.get_setting("openai_api_key", "").strip()
        if not api_key:
            self.append_chat_message("Sistema", "Erro: OpenAI API Key não configurada nas Configurações Globais.")
            self.playing = False
            self.execute_button.config(state=tk.NORMAL)
            if hasattr(self, 'stop_execution_button'):
                self.stop_execution_button.config(state=tk.DISABLED)
            return

        self.agent_active = True
        self.agent_status_var.set("Status: Pensando...")
        self.append_chat_message("Agente", f"Iniciando tarefa [{ai_mode}]: '{task_description}'")

        import requests
        from PIL import ImageGrab
        import io
        import base64
        from pynput.mouse import Controller as MouseController, Button as MouseButton
        from pynput.keyboard import Controller as KeyController

        mouse_ctrl = MouseController()
        key_ctrl = KeyController()
        executed_actions = []

        # Lê o conteúdo do arquivo 1 se existir
        file_content_1 = ""
        if file_path and os.path.exists(file_path):
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                file_content_1 = f.read()

        # Para Selenium, valida se o navegador está ativo
        if ai_mode == "selenium":
            if not SELENIUM_AVAILABLE or not hasattr(self, 'selenium_driver') or not self.selenium_driver:
                self.append_chat_message("Sistema", "Erro: O navegador Selenium não está ativo. Por favor, abra-o primeiro.")
                self.agent_active = False
                self.playing = False
                self.agent_status_var.set("Status: Inativo")
                self.execute_button.config(state=tk.NORMAL)
                if hasattr(self, 'stop_execution_button'):
                    self.stop_execution_button.config(state=tk.DISABLED)
                return

        # Lógica de Mesclagem de Arquivos Excel (excel_merge)
        if ai_mode == "excel_merge":
            self.agent_status_var.set("Status: Mesclando arquivos...")
            try:
                file_content_2 = ""
                if file_path_2 and os.path.exists(file_path_2):
                    with open(file_path_2, 'r', encoding='utf-8', errors='ignore') as f:
                        file_content_2 = f.read()
                
                file_content_3 = ""
                if file_path_3 and os.path.exists(file_path_3):
                    with open(file_path_3, 'r', encoding='utf-8', errors='ignore') as f:
                        file_content_3 = f.read()

                prompt = (
                    f"Você é um engenheiro de dados especialista. O usuário quer realizar a seguinte mesclagem/consolidação:\n"
                    f"'{task_description}'.\n\n"
                    f"Temos os conteúdos dos seguintes arquivos de entrada:\n"
                    f"=== ARQUIVO 1 ===\n{file_content_1}\n==================\n\n"
                    f"=== ARQUIVO 2 ===\n{file_content_2}\n==================\n\n"
                    f"=== ARQUIVO 3 ===\n{file_content_3}\n==================\n\n"
                    f"Por favor, consolide e formate profissionalmente os dados desses arquivos em uma única tabela estruturada.\n"
                    f"Retorne obrigatoriamente um objeto JSON contendo a lista de linhas sob a chave 'rows', onde cada linha é um dicionário chave-valor (cabeçalho da coluna -> valor).\n"
                    f"Exemplo:\n"
                    f"{{\n"
                    f"  \"rows\": [\n"
                    f"    {{\"Nome\": \"Willian\", \"Email\": \"willian@teste.com\", \"Telefone\": \"99999-9999\", \"Origem\": \"Arquivo 1\"}},\n"
                    f"    {{\"Nome\": \"Danilo\", \"Email\": \"danilo@teste.com\", \"Telefone\": \"88888-8888\", \"Origem\": \"Arquivo 2\"}}\n"
                    f"  ]\n"
                    f"}}"
                )
                
                headers = {
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json"
                }
                
                payload = {
                    "model": "gpt-4o",
                    "response_format": {"type": "json_object"},
                    "messages": [
                        {"role": "user", "content": prompt}
                    ],
                    "max_tokens": 3000
                }
                
                res = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, json=payload)
                res.raise_for_status()
                
                res_data = res.json()
                ai_message = res_data['choices'][0]['message']['content'].strip()
                
                if ai_message.startswith("```json"):
                    ai_message = ai_message[7:]
                if ai_message.endswith("```"):
                    ai_message = ai_message[:-3]
                ai_message = ai_message.strip()
                
                # Gera a planilha Excel formatada profissionalmente
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter
                
                data_json = json.loads(ai_message)
                rows = data_json.get("rows", [])
                
                if not rows:
                    raise ValueError("Nenhum dado retornado pela IA para mesclagem.")
                    
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Consolidado"
                
                # Cabeçalhos
                headers_list = list(rows[0].keys())
                for col_idx, h_text in enumerate(headers_list, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=h_text)
                    cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                # Dados
                thin_border = Border(
                    left=Side(style='thin', color='D3D3D3'),
                    right=Side(style='thin', color='D3D3D3'),
                    top=Side(style='thin', color='D3D3D3'),
                    bottom=Side(style='thin', color='D3D3D3')
                )
                
                for row_idx, r_data in enumerate(rows, start=2):
                    for col_idx, h_text in enumerate(headers_list, start=1):
                        val = r_data.get(h_text, "")
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.font = Font(name="Arial", size=10)
                        cell.alignment = Alignment(vertical="center")
                        cell.border = thin_border
                        
                        # Estilo zebrado
                        if row_idx % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # Auto-ajuste de colunas
                for col in ws.columns:
                    max_len = 0
                    for cell in col:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                    
                # Salva o arquivo
                out_p = output_path or "consolidado.xlsx"
                wb.save(out_p)
                self.append_chat_message("Agente", f"Consolidação concluída com sucesso em: '{out_p}'!")
            except Exception as e:
                self.append_chat_message("Sistema", f"Erro crítico na mesclagem Excel: {str(e)}")
            finally:
                self.agent_active = False
                self.playing = False
                self.agent_status_var.set("Status: Inativo")
                self.execute_button.config(state=tk.NORMAL)
                if hasattr(self, 'stop_execution_button'):
                    self.stop_execution_button.config(state=tk.DISABLED)
            return

        # Loops de Execução Autônoma para OCR e Selenium
        max_iterations = 12
        iteration = 0
        success = False

        try:
            while iteration < max_iterations and self.playing:
                iteration += 1
                self.agent_status_var.set(f"Status: Executando {ai_mode} ({iteration}/{max_iterations})...")

                # Configura requisições para a API OpenAI
                headers = {
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json"
                }

                recent_chat = self.chat_history_list[-8:] if self.chat_history_list else []
                chat_context_str = "\n".join([f"[{m['role']}]: {m['content']}" for m in recent_chat])

                if ai_mode == "ocr":
                    # Captura a tela física
                    try:
                        user32 = ctypes.windll.user32
                        screen_w = user32.GetSystemMetrics(0)
                        screen_h = user32.GetSystemMetrics(1)
                    except Exception:
                        screen_w, screen_h = 1920, 1080

                    screenshot = ImageGrab.grab()
                    img_w, img_h = screenshot.size

                    buffer = io.BytesIO()
                    screenshot.save(buffer, format="JPEG", quality=80)
                    img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

                    prompt = (
                        f"Você é um robô de RPA inteligente executando em loop autônomo. O usuário quer que você execute a seguinte tarefa: '{task_description}'.\n"
                        f"Temos dados de um arquivo local que podem ser relevantes:\n"
                        f"=== DADOS DO ARQUIVO ===\n{file_content_1}\n========================\n\n"
                        f"A resolução física do monitor atual é de {screen_w}x{screen_h} pixels.\n"
                        f"A imagem enviada é um print da tela inteira do monitor. A resolução do print é {img_w}x{img_h}.\n"
                        f"Aqui está o histórico recente do chat:\n"
                        f"=== HISTÓRICO DE CHAT ===\n{chat_context_str}\n========================\n\n"
                        f"Aqui está a lista de ações já executadas para evitar repetições:\n"
                        f"{json.dumps(executed_actions, indent=2)}\n\n"
                        f"Você deve identificar o estado atual da tela e o próximo passo para realizar a tarefa.\n"
                        f"Retorne OBRIGATORIAMENTE um objeto JSON puro contendo a resposta. Não use blocos de marcação adicionais de código (como ```json) ou explicações. O JSON deve seguir exatamente o seguinte formato:\n"
                        f"{{\n"
                        f"  \"completed\": false, // defina como true apenas quando verificar na tela que o objetivo final da tarefa foi totalmente alcançado\n"
                        f"  \"message\": \"Texto em português descrevendo sua análise do estado atual da tela e o que vai fazer agora, ou pergunta para o usuário\",\n"
                        f"  \"waiting_for_user\": false, // defina como true se você precisar de alguma resposta ou ajuda do usuário no chat para prosseguir\n"
                        f"  \"actions\": [\n"
                        f"    // Uma ou mais ações a serem executadas sequencialmente nesta iteração. Exemplos:\n"
                        f"    // {{\"type\": \"click\", \"x\": 150, \"y\": 300}},\n"
                        f"    // {{\"type\": \"type\", \"text\": \"valor a digitar\"}},\n"
                        f"    // {{\"type\": \"scroll\", \"dx\": 0, \"dy\": -100}},\n"
                        f"    // {{\"type\": \"wait\", \"duration\": 2.0}}\n"
                        f"  ]\n"
                        f"}}"
                    )

                    payload = {
                        "model": "gpt-4o",
                        "response_format": {"type": "json_object"},
                        "messages": [
                            {
                                "role": "user",
                                "content": [
                                    {"type": "text", "text": prompt},
                                    {
                                        "type": "image_url",
                                        "image_url": {
                                            "url": f"data:image/jpeg;base64,{img_base64}"
                                        }
                                    }
                                ]
                            }
                        ],
                        "max_tokens": 1000,
                        "temperature": 0.2
                    }

                elif ai_mode == "selenium":
                    driver = self.selenium_driver
                    elements_info = []
                    for el in driver.find_elements(By.XPATH, "//*[self::input or self::textarea or self::select or self::button]"):
                        try:
                            if el.is_displayed() and el.is_enabled():
                                elements_info.append({
                                    "tag": el.tag_name,
                                    "id": el.get_attribute("id") or "",
                                    "name": el.get_attribute("name") or "",
                                    "placeholder": el.get_attribute("placeholder") or "",
                                    "type": el.get_attribute("type") or "",
                                    "text": el.text or ""
                                })
                        except Exception:
                            pass

                    prompt = (
                        f"Você é um assistente de automação web Selenium executando em loop autônomo. O usuário quer executar a tarefa: '{task_description}'.\n"
                        f"Temos dados de um arquivo local relevante:\n"
                        f"=== DADOS DO ARQUIVO ===\n{file_content_1}\n========================\n\n"
                        f"Aqui está a lista de elementos interativos encontrados na página HTML atual:\n"
                        f"{json.dumps(elements_info, indent=2)}\n\n"
                        f"Aqui está o histórico recente do chat:\n"
                        f"=== HISTÓRICO DE CHAT ===\n{chat_context_str}\n========================\n\n"
                        f"Aqui está a lista de ações já executadas para evitar repetições:\n"
                        f"{json.dumps(executed_actions, indent=2)}\n\n"
                        f"Retorne obrigatoriamente um objeto JSON puro. O JSON deve listar as ações informando o seletor correspondente (preferencialmente por 'id' se houver, ou 'name', ou 'placeholder'). Formato:\n"
                        f"{{\n"
                        f"  \"completed\": false, // defina como true quando verificar que a tarefa web foi concluída\n"
                        f"  \"message\": \"Texto em português informando o que está fazendo ou perguntando algo ao usuário\",\n"
                        f"  \"waiting_for_user\": false, // se precisar de resposta do usuário no chat antes do próximo passo\n"
                        f"  \"actions\": [\n"
                        f"    // Exemplos:\n"
                        f"    // {{\"type\": \"click\", \"by\": \"id\", \"value\": \"id_do_elemento\"}},\n"
                        f"    // {{\"type\": \"type\", \"by\": \"id\", \"value\": \"id_do_elemento\", \"text\": \"valor\"}},\n"
                        f"    // {{\"type\": \"wait\", \"duration\": 1.0}}\n"
                        f"  ]\n"
                        f"}}"
                    )

                    payload = {
                        "model": "gpt-4o-mini",
                        "response_format": {"type": "json_object"},
                        "messages": [
                            {"role": "user", "content": prompt}
                        ],
                        "max_tokens": 1000,
                        "temperature": 0.2
                    }

                res = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, json=payload)
                res.raise_for_status()
                res_data = res.json()
                ai_message = res_data['choices'][0]['message']['content'].strip()

                if ai_message.startswith("```json"):
                    ai_message = ai_message[7:]
                if ai_message.endswith("```"):
                    ai_message = ai_message[:-3]
                ai_message = ai_message.strip()

                response_json = json.loads(ai_message)

                msg_text = response_json.get("message", "")
                if msg_text:
                    self.append_chat_message("Agente", msg_text)

                if response_json.get("completed", False):
                    success = True
                    self.append_chat_message("Agente", "Objetivo alcançado! Tarefa concluída.")
                    break

                if response_json.get("waiting_for_user", False):
                    self.user_message_event.clear()
                    self.agent_status_var.set("Status: Aguardando resposta do usuário...")
                    while self.playing and not self.user_message_event.is_set():
                        time.sleep(0.2)
                    
                    if not self.playing:
                        break
                    continue

                # Executa ações
                actions_list = response_json.get("actions", [])
                for act in actions_list:
                    if not self.playing:
                        break

                    act_type = act.get("type")
                    if ai_mode == "ocr":
                        if act_type == "click":
                            x, y = act.get("x"), act.get("y")
                            mouse_ctrl.position = (x, y)
                            time.sleep(0.2)
                            mouse_ctrl.press(MouseButton.left)
                            time.sleep(0.05)
                            mouse_ctrl.release(MouseButton.left)
                            time.sleep(0.5)
                            executed_actions.append(f"Clique em ({x}, {y})")

                        elif act_type == "type":
                            text = act.get("text")
                            key_ctrl.type(text)
                            time.sleep(0.2)
                            executed_actions.append(f"Digitação de '{text}'")

                        elif act_type == "scroll":
                            dx, dy = act.get("dx", 0), act.get("dy", 0)
                            mouse_ctrl.scroll(dx, dy)
                            time.sleep(0.3)
                            executed_actions.append(f"Rolagem (dx={dx}, dy={dy})")

                        elif act_type == "wait":
                            duration = act.get("duration", 1.0)
                            time.sleep(duration)
                            executed_actions.append(f"Espera de {duration}s")
                            
                    elif ai_mode == "selenium":
                        driver = self.selenium_driver
                        by_type = act.get("by")
                        val = act.get("value")
                        element = None
                        try:
                            if by_type == "id":
                                element = driver.find_element(By.ID, val)
                            elif by_type == "name":
                                element = driver.find_element(By.NAME, val)
                            elif by_type == "placeholder":
                                element = driver.find_element(By.XPATH, f"//input[@placeholder='{val}'] | //textarea[@placeholder='{val}']")
                        except Exception:
                            pass

                        if element:
                            if act_type == "click":
                                element.click()
                                time.sleep(0.5)
                                executed_actions.append(f"Clique Selenium no elemento '{val}'")
                            elif act_type == "type":
                                element.clear()
                                element.send_keys(act["text"])
                                time.sleep(0.3)
                                executed_actions.append(f"Digitação Selenium no elemento '{val}' de '{act['text']}'")
                            elif act_type == "wait":
                                duration = act.get("duration", 1.0)
                                time.sleep(duration)
                                executed_actions.append(f"Espera Selenium de {duration}s")
                        else:
                            executed_actions.append(f"Elemento '{val}' não localizado nesta iteração.")

                time.sleep(1.0)

            if not success and iteration >= max_iterations:
                self.append_chat_message("Agente", "Limite máximo de iterações atingido sem confirmação de sucesso.")

        except Exception as e:
            self.append_chat_message("Sistema", f"Erro crítico na execução do agente: {str(e)}")
            print(f"Erro no agente autônomo: {e}")
        finally:
            self.agent_active = False
            self.playing = False
            self.agent_status_var.set("Status: Inativo")
            self.execute_button.config(state=tk.NORMAL)
            if hasattr(self, 'stop_execution_button'):
                self.stop_execution_button.config(state=tk.DISABLED)

    def open_selenium(self):
        """Abre uma janela de automação web com Selenium."""
        if not SELENIUM_AVAILABLE:
            messagebox.showerror("Selenium não disponível", "Selenium não está instalado.")
            return

        selenium_window = tk.Toplevel(self.root)
        selenium_window.title("Automação Web com Selenium")
        selenium_window.geometry("480x350")
        selenium_window.resizable(False, False)

        tk.Label(selenium_window, text="Automação de Navegador", font=("Arial", 12, "bold")).pack(anchor=tk.W, padx=10, pady=(10, 5))

        tk.Label(selenium_window, text="URL:", font=("Arial", 10, "bold")).pack(anchor=tk.W, padx=10)
        url_var = tk.StringVar(value="https://www.google.com")
        url_entry = tk.Entry(selenium_window, textvariable=url_var, width=50)
        url_entry.pack(padx=10, pady=(0, 10))

        tk.Label(selenium_window, text="Ações disponíveis:", font=("Arial", 10, "bold")).pack(anchor=tk.W, padx=10, pady=(10, 5))

        actions_frame = tk.Frame(selenium_window)
        actions_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        def open_browser():
            try:
                url = url_var.get().strip() or "https://www.google.com"
                self.selenium_driver = webdriver.Chrome()
                self.selenium_driver.get(url)
                messagebox.showinfo("Sucesso", f"Navegador aberto em: {url}")
                status_label.config(text="Status: Navegador ativo")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao abrir navegador: {e}")

        def close_browser():
            try:
                if hasattr(self, 'selenium_driver') and self.selenium_driver:
                    self.selenium_driver.quit()
                    status_label.config(text="Status: Navegador fechado")
                    messagebox.showinfo("Sucesso", "Navegador fechado.")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao fechar navegador: {e}")

        def find_element():
            try:
                selector = simpledialog.askstring("Buscar elemento", "Digite o seletor CSS:")
                if selector and hasattr(self, 'selenium_driver') and self.selenium_driver:
                    element = self.selenium_driver.find_element(By.CSS_SELECTOR, selector)
                    element.click()
                    messagebox.showinfo("Sucesso", f"Elemento clicado: {selector}")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao clicar elemento: {e}")

        def fill_form():
            try:
                selector = simpledialog.askstring("Preencher campo", "Digite o seletor CSS do campo:")
                if selector and hasattr(self, 'selenium_driver') and self.selenium_driver:
                    text = simpledialog.askstring("Texto", "Digite o texto a preencher:")
                    if text:
                        element = self.selenium_driver.find_element(By.CSS_SELECTOR, selector)
                        element.clear()
                        element.send_keys(text)
                        messagebox.showinfo("Sucesso", f"Campo preenchido: {text}")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao preencher campo: {e}")

        tk.Button(actions_frame, text="Abrir Navegador", command=open_browser, width=20, bg="#2196f3", fg="white").pack(pady=5)
        tk.Button(actions_frame, text="Clicar em Elemento", command=find_element, width=20, bg="#4caf50", fg="white").pack(pady=5)
        tk.Button(actions_frame, text="Preencher Campo", command=fill_form, width=20, bg="#ff9800", fg="white").pack(pady=5)
        tk.Button(actions_frame, text="Fechar Navegador", command=close_browser, width=20, bg="#f44336", fg="white").pack(pady=5)

        status_label = tk.Label(selenium_window, text="Status: Aguardando ação...", font=("Arial", 9), fg="#666")
        status_label.pack(side=tk.BOTTOM, padx=10, pady=10)


if __name__ == "__main__":
    root = tk.Tk()
    app = ClickRecorderApp(root)
    root.mainloop()
